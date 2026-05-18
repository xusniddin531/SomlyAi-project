"""
Excel report generation service for Somly AI.
Generates financial reports with Summary, Transactions, and Debts sheets.
"""

import os
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from src.database import (
    get_user,
    get_transactions_paginated,
    get_user_all_balances,
    debts_collection,
)


def _safe_filename_part(s: str) -> str:
    """Fayl nomi uchun xavfsiz qism: maxsus belgilar olib tashlanadi."""
    if not s:
        return "User"
    # Faqat harflar/raqamlar/probel/tire/_ qoldiramiz
    cleaned = re.sub(r'[^\w\s\-]', '', str(s), flags=re.UNICODE)
    cleaned = re.sub(r'\s+', '_', cleaned.strip())
    return cleaned[:40] or "User"



# Color schemes
COLOR_DARK_BLUE = "1E3A8A"
COLOR_WHITE = "FFFFFF"
COLOR_GREEN = "D1FAE5"
COLOR_RED = "FECACA"
COLOR_YELLOW = "FEF3C7"
COLOR_LIGHT_BLUE = "DBEAFE"
COLOR_GRAY = "F3F4F6"
COLOR_INCOME_TEXT = "059669"
COLOR_EXPENSE_TEXT = "DC2626"
COLOR_HEADER_BG = "1E40AF"


async def generate_excel_report(
    telegram_id: int,
    date_from: datetime,
    date_to: datetime,
    language: str = "uz",
) -> str:
    """
    Generate Excel report with three sheets: Summary, Transactions, Debts.
    Returns file path.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Get user data
    user = await get_user(telegram_id)
    full_name = user.get("full_name", "Foydalanuvchi")

    # Create sheets
    await _create_summary_sheet(wb, telegram_id, date_from, date_to, full_name, language)
    await _create_transactions_sheet(
        wb, telegram_id, date_from, date_to, language
    )
    await _create_debts_sheet(wb, telegram_id, language)

    # Save file with Uzbek month name
    months_uz = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
        7: "Iyul", 8: "Avgust", 9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr"
    }
    month_uz = months_uz.get(date_from.month, str(date_from.month))
    safe_name = _safe_filename_part(full_name)
    filename = f"Somly_{safe_name}_{month_uz}-{date_from.year}.xlsx"

    # temp/ papkasini yaratamiz (idempotent — mavjud bo'lsa skip)
    os.makedirs("temp", exist_ok=True)
    filepath = os.path.join("temp", filename)
    wb.save(filepath)
    return filepath


async def _create_summary_sheet(
    wb: Workbook,
    telegram_id: int,
    date_from: datetime,
    date_to: datetime,
    full_name: str,
    language: str,
) -> None:
    """Create the summary/overview sheet."""
    ws = wb.create_sheet("Xulosa", 0)
    
    # Get user balances
    all_balances = await get_user_all_balances(telegram_id)
    
    # Get transactions for the period (use date strings for filtering)
    main_currency = "UZS" if "UZS" in all_balances else list(all_balances.keys())[0]
    
    date_from_str = date_from.strftime("%Y-%m-%d")
    date_to_str = date_to.strftime("%Y-%m-%d")
    
    txs, _ = await get_transactions_paginated(
        telegram_id,
        page=1,
        per_page=10000,
        date_from=date_from_str,
        date_to=date_to_str,
    )
    
    # Calculate totals by type and currency
    income_by_currency = {}
    expense_by_currency = {}
    category_totals = {}
    
    for tx in txs:
        currency = tx.get("currency", "UZS")
        tx_type = tx.get("type", "")
        amount = tx.get("amount", 0)
        category = tx.get("category", "Uncategorized")
        
        if tx_type == "kirim":
            income_by_currency[currency] = income_by_currency.get(currency, 0) + amount
        elif tx_type == "chiqim":
            expense_by_currency[currency] = expense_by_currency.get(currency, 0) + amount
        
        if tx_type in ["kirim", "chiqim"]:
            category_totals[category] = category_totals.get(category, {})
            category_totals[category][tx_type] = category_totals[category].get(tx_type, 0) + amount
    
    # ════════════════════════════════════════
    # Header section
    # ════════════════════════════════════════
    row = 1
    
    # Title
    title_cell = ws[f"A{row}"]
    title_cell.value = "Somly AI — Moliyaviy Hisobot"
    title_cell.font = Font(name="Arial", size=16, bold=True, color=COLOR_HEADER_BG)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    
    # User
    ws[f"A{row}"].value = f"Foydalanuvchi: {full_name}"
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    
    # Period
    period_text = f"Davr: {date_from.strftime('%d %B %Y').replace('April', 'Aprel').replace('May', 'May').replace('June', 'Iyun').replace('January', 'Yanvar').replace('February', 'Fevral').replace('March', 'Mart').replace('July', 'Iyul').replace('August', 'Avgust').replace('September', 'Sentyabr').replace('October', 'Oktyabr').replace('November', 'Noyabr').replace('December', 'Dekabr')} — {date_to.strftime('%d %B %Y').replace('April', 'Aprel').replace('May', 'May').replace('June', 'Iyun').replace('January', 'Yanvar').replace('February', 'Fevral').replace('March', 'Mart').replace('July', 'Iyul').replace('August', 'Avgust').replace('September', 'Sentyabr').replace('October', 'Oktyabr').replace('November', 'Noyabr').replace('December', 'Dekabr')}"
    ws[f"A{row}"].value = period_text
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    
    # Created date
    created_text = f"Yaratildi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws[f"A{row}"].value = created_text
    ws.merge_cells(f"A{row}:F{row}")
    row += 2
    
    # ════════════════════════════════════════
    # Key metrics
    # ════════════════════════════════════════
    total_income = income_by_currency.get(main_currency, 0)
    total_expense = expense_by_currency.get(main_currency, 0)
    net_balance = total_income - total_expense
    
    # Income
    ws[f"A{row}"].value = "Jami kirim:"
    ws[f"A{row}"].font = Font(size=12, bold=True)
    income_cell = ws[f"B{row}"]
    income_cell.value = f"{total_income:,.0f} {main_currency}"
    income_cell.font = Font(size=12, bold=True, color="FFFFFF")
    income_cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    income_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"B{row}:C{row}")
    row += 1
    
    # Expense
    ws[f"A{row}"].value = "Jami chiqim:"
    ws[f"A{row}"].font = Font(size=12, bold=True)
    expense_cell = ws[f"B{row}"]
    expense_cell.value = f"{total_expense:,.0f} {main_currency}"
    expense_cell.font = Font(size=12, bold=True, color="FFFFFF")
    expense_cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    expense_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"B{row}:C{row}")
    row += 1
    
    # Net balance
    ws[f"A{row}"].value = "Sof qoldiq:"
    ws[f"A{row}"].font = Font(size=12, bold=True)
    balance_cell = ws[f"B{row}"]
    balance_cell.value = f"{net_balance:,.0f} {main_currency}"
    balance_cell.font = Font(size=12, bold=True, color="FFFFFF")
    balance_cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    balance_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"B{row}:C{row}")
    row += 2
    
    # ════════════════════════════════════════
    # Categories breakdown table
    # ════════════════════════════════════════
    ws[f"A{row}"].value = "KATEGORIYALAR"
    ws[f"A{row}"].font = Font(size=11, bold=True, color=COLOR_WHITE)
    ws[f"A{row}"].fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    
    # Headers
    headers = ["Kategoriya", "Kirim", "Chiqim", "Foiz (%)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    # Sort categories by total spending (descending)
    sorted_categories = sorted(
        category_totals.items(),
        key=lambda x: x[1].get("chiqim", 0) + x[1].get("kirim", 0),
        reverse=True
    )
    
    for category, amounts in sorted_categories:
        income = amounts.get("kirim", 0)
        expense = amounts.get("chiqim", 0)
        total = income + expense
        percentage = (expense / total_expense * 100) if total_expense > 0 else 0
        
        ws.cell(row=row, column=1).value = category
        ws.cell(row=row, column=2).value = income
        ws.cell(row=row, column=3).value = expense
        ws.cell(row=row, column=4).value = percentage
        
        # Format
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).number_format = '0.0"%"'
        
        # Alternating colors
        if row % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
        
        row += 1
    
    # ════════════════════════════════════════
    # Debts summary
    # ════════════════════════════════════════
    row += 1
    debts = await debts_collection.find({"telegram_id": telegram_id, "status": "active"}).to_list(None)
    
    to_give = sum(d.get("amount", 0) - d.get("paid_amount", 0) for d in debts if d.get("direction") == "bergan")
    to_receive = sum(d.get("amount", 0) - d.get("paid_amount", 0) for d in debts if d.get("direction") == "olgan")
    
    ws[f"A{row}"].value = "QARZLAR"
    ws[f"A{row}"].font = Font(size=11, bold=True, color=COLOR_WHITE)
    ws[f"A{row}"].fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    
    ws[f"A{row}"].value = "Berishim kerak:"
    ws[f"B{row}"].value = f"{to_give:,.0f}"
    ws[f"B{row}"].font = Font(bold=True, color=COLOR_EXPENSE_TEXT)
    row += 1
    
    ws[f"A{row}"].value = "Olishim kerak:"
    ws[f"B{row}"].value = f"{to_receive:,.0f}"
    ws[f"B{row}"].font = Font(bold=True, color=COLOR_INCOME_TEXT)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15


async def _create_transactions_sheet(
    wb: Workbook,
    telegram_id: int,
    date_from: datetime,
    date_to: datetime,
    language: str,
) -> None:
    """Create transactions sheet with daily grouping."""
    ws = wb.create_sheet("Tranzaksiyalar", 1)
    
    # Get all transactions for period (use date strings for filtering)
    date_from_str = date_from.strftime("%Y-%m-%d")
    date_to_str = date_to.strftime("%Y-%m-%d")
    
    txs, _ = await get_transactions_paginated(
        telegram_id,
        page=1,
        per_page=10000,
        date_from=date_from_str,
        date_to=date_to_str,
    )
    
    # Sort by date
    txs = sorted(txs, key=lambda x: x.get("date", ""))
    
    # Headers
    headers = ["№", "Sana", "Kun", "Tur", "Summa", "Valyuta", "Kategoriya", "Izoh", "Balans (keyin)", "Shaxs"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.freeze_panes = "A2"
    
    row = 2
    tx_counter = 1
    current_date = None
    daily_income = 0
    daily_expense = 0
    opening_balance = 0
    
    for tx in txs:
        tx_date = tx.get("date", "")
        if not tx_date:
            continue
        
        # Parse date
        try:
            tx_datetime = datetime.fromisoformat(tx_date.split("T")[0])
        except:
            continue
        
        # Check if new day
        if tx_datetime != current_date:
            if current_date is not None:
                # Add day summary row
                ws[f"A{row}"].value = f"Kun yakuni: Kirim {daily_income:,.0f} | Chiqim {daily_expense:,.0f} | Qoldiq {opening_balance + daily_income - daily_expense:,.0f}"
                ws.merge_cells(f"A{row}:J{row}")
                ws[f"A{row}"].font = Font(italic=True, size=9)
                ws[f"A{row}"].fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
                row += 1
                daily_income = 0
                daily_expense = 0
            
            # Add date group header
            date_str = tx_datetime.strftime("%Y-%yil, %d-%B, %A").replace("04", "Aprel").replace("05", "May")
            ws[f"A{row}"].value = date_str
            ws.merge_cells(f"A{row}:J{row}")
            ws[f"A{row}"].font = Font(bold=True, size=10)
            ws[f"A{row}"].fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
            row += 1
            
            # Add opening balance row
            ws[f"A{row}"].value = f"Kun boshiga qoldiq: {opening_balance:,.0f} UZS"
            ws.merge_cells(f"A{row}:J{row}")
            ws[f"A{row}"].fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
            row += 1
            
            current_date = tx_datetime
        
        # Transaction row
        ws[f"A{row}"].value = tx_counter
        ws[f"B{row}"].value = tx_date.split("T")[0]
        ws[f"C{row}"].value = tx_datetime.strftime("%A").replace("Monday", "Dushanba").replace("Tuesday", "Seshanba").replace("Wednesday", "Chorshanba").replace("Thursday", "Payshanba").replace("Friday", "Juma").replace("Saturday", "Shanba").replace("Sunday", "Yakshanba")
        ws[f"D{row}"].value = "Kirim" if tx.get("type") == "kirim" else "Chiqim"
        ws[f"E{row}"].value = tx.get("amount", 0)
        ws[f"F{row}"].value = tx.get("currency", "UZS")
        ws[f"G{row}"].value = tx.get("category", "")
        ws[f"H{row}"].value = tx.get("description", "")
        ws[f"I{row}"].value = opening_balance + (tx.get("amount", 0) if tx.get("type") == "kirim" else -tx.get("amount", 0))
        ws[f"J{row}"].value = tx.get("person", "")
        
        # Formatting
        ws[f"E{row}"].number_format = '#,##0'
        ws[f"I{row}"].number_format = '#,##0'
        
        # Color by type
        if tx.get("type") == "kirim":
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
            daily_income += tx.get("amount", 0)
        else:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")
            daily_expense += tx.get("amount", 0)
        
        opening_balance += (tx.get("amount", 0) if tx.get("type") == "kirim" else -tx.get("amount", 0))
        tx_counter += 1
        row += 1
    
    # Final day summary
    if current_date is not None:
        ws[f"A{row}"].value = f"Kun yakuni: Kirim {daily_income:,.0f} | Chiqim {daily_expense:,.0f} | Qoldiq {opening_balance:,.0f}"
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"].font = Font(italic=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
    
    # Adjust column widths
    widths = [5, 12, 12, 10, 12, 10, 15, 20, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


async def _create_debts_sheet(
    wb: Workbook,
    telegram_id: int,
    language: str,
) -> None:
    """Create debts sheet."""
    ws = wb.create_sheet("Qarzlar", 2)
    
    # Get debts
    debts = await debts_collection.find({"telegram_id": telegram_id}).to_list(None)
    
    row = 1
    
    # ════════════════════════════════════════
    # To give section (Berishim kerak)
    # ════════════════════════════════════════
    ws[f"A{row}"].value = "BERISHIM KERAK (men qarzman)"
    ws[f"A{row}"].font = Font(size=11, bold=True, color=COLOR_WHITE)
    ws[f"A{row}"].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    
    # Headers for to-give section
    headers = ["Ism", "Summa", "Valyuta", "Sana", "Muddat", "Holat"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    to_give_debts = [d for d in debts if d.get("direction") == "bergan"]
    for debt in to_give_debts:
        ws[f"A{row}"].value = debt.get("person", "")
        ws[f"B{row}"].value = debt.get("amount", 0) - debt.get("paid_amount", 0)
        ws[f"C{row}"].value = debt.get("currency", "UZS")
        ws[f"D{row}"].value = debt.get("date", "")
        ws[f"E{row}"].value = debt.get("due_date", "")
        
        status_map = {"active": "Aktiv", "paid": "To'landi", "partial": "Qisman", "cancelled": "Bekor"}
        ws[f"F{row}"].value = status_map.get(debt.get("status", "active"), "Noma'lum")
        
        ws[f"B{row}"].number_format = '#,##0'
        
        # Alternating colors
        if row % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
        
        row += 1
    
    row += 1
    
    # ════════════════════════════════════════
    # To receive section (Olishim kerak)
    # ════════════════════════════════════════
    ws[f"A{row}"].value = "OLISHIM KERAK (u qarzdir)"
    ws[f"A{row}"].font = Font(size=11, bold=True, color=COLOR_WHITE)
    ws[f"A{row}"].fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    
    # Headers for to-receive section
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    to_receive_debts = [d for d in debts if d.get("direction") == "olgan"]
    for debt in to_receive_debts:
        ws[f"A{row}"].value = debt.get("person", "")
        ws[f"B{row}"].value = debt.get("amount", 0) - debt.get("paid_amount", 0)
        ws[f"C{row}"].value = debt.get("currency", "UZS")
        ws[f"D{row}"].value = debt.get("date", "")
        ws[f"E{row}"].value = debt.get("due_date", "")
        
        status_map = {"active": "Aktiv", "paid": "To'landi", "partial": "Qisman", "cancelled": "Bekor"}
        ws[f"F{row}"].value = status_map.get(debt.get("status", "active"), "Noma'lum")
        
        ws[f"B{row}"].number_format = '#,##0'
        
        # Alternating colors
        if row % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
